from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment


def align(sheet, start_row, end_row, start_col, end_col, horizontal, vertical):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            sheet.cell(row = row, column = col).alignment = Alignment(horizontal=horizontal, vertical = vertical)

def table_head(sheet, start_row, end_row, start_col, end_col):
    fill = PatternFill(start_color='212121', end_color='212121', fill_type='solid')
    font = Font(color = 'FFFFFF', bold = True) #, size = 20)
    for row in range(start_row, end_row+1):
        for col in range(start_col, end_col+1):
            sheet.cell(row = row, column = col).fill = fill
            sheet.cell(row = row, column = col).font = font

def column_head(sheet, start_row, end_row, start_col, end_col):
    clr = '212121'
    fill = PatternFill(start_color='B3E5FC', end_color='B3E5FC', fill_type='solid')
    font = Font(color = '212121', bold = True) #, size = 20)
    border = Border(left=Side(style='thin', color = '212121'), right=Side(style='thin', color = '212121'), top=Side(style='thick',color = '212121'), bottom=Side(style='thick',color = '212121'))
    for row in range(start_row, end_row+1):
        for col in range(start_col, end_col+1):
            sheet.cell(row = row, column = col).fill = fill
            sheet.cell(row = row, column = col).font = font
            sheet.cell(row = row, column = col).border = border

def column_rows(sheet, start_row, end_row, start_col, end_col):
    clr = '212121'
    #fill = PatternFill(start_color='B3E5FC', end_color='B3E5FC', fill_type='solid')
    #font = Font(color = '212121', bold = True) #, size = 20)
    thick_border_last = Border(left=Side(style='thin', color = '212121'), right=Side(style='thin', color = '212121'), top=Side(style='thin',color = '212121'), bottom=Side(style='thick',color = '212121'))
    thick_border_first = Border(left=Side(style='thin', color = '212121'), right=Side(style='thin', color = '212121'), top=Side(style='thick',color = '212121'), bottom=Side(style='thin',color = '212121'))
    thick_border_both = Border(left=Side(style='thin', color = '212121'), right=Side(style='thin', color = '212121'), top=Side(style='thick',color = '212121'), bottom=Side(style='thick',color = '212121'))
    thin_border = Border(left=Side(style='thin', color = '212121'), right=Side(style='thin', color = '212121'), top=Side(style='thin',color = '212121'), bottom=Side(style='thin',color = '212121'))
    for row in range(start_row, end_row+1):
        if row == start_row:
            if end_row == start_row:
                for col in range(start_col, end_col+1):
                    #sheet.cell(row = row, column = col).fill = fill
                    #sheet.cell(row = row, column = col).font = font
                    sheet.cell(row = row, column = col).border = thick_border_both

            else:
                for col in range(start_col, end_col+1):
                    #sheet.cell(row = row, column = col).fill = fill
                    #sheet.cell(row = row, column = col).font = font
                    sheet.cell(row = row, column = col).border = thick_border_first

        elif row == end_row:
            for col in range(start_col, end_col+1):
                #sheet.cell(row = row, column = col).fill = fill
                #sheet.cell(row = row, column = col).font = font
                sheet.cell(row = row, column = col).border = thick_border_last

        else:
            for col in range(start_col, end_col+1):
                #sheet.cell(row = row, column = col).fill = fill
                #sheet.cell(row = row, column = col).font = font
                sheet.cell(row = row, column = col).border = thin_border

def table_foot(sheet, start_row, end_row, start_col, end_col):
    clr = '212121'
    fill = PatternFill(start_color='FFF9C4', end_color='212121', fill_type='solid')
    font = Font(color = '212121', bold = True, size = 10)
    thick_border_last = Border(left=Side(style='thin', color = '212121'), right=Side(style='thin', color = '212121'), top=Side(style='thin',color = '212121'), bottom=Side(style='thick',color = '212121'))
    for row in range(start_row, end_row+1):
        for col in range(start_col, end_col+1):
            sheet.cell(row = row, column = col).fill = fill
            sheet.cell(row = row, column = col).font = font
            sheet.cell(row = row, column = col).border = thick_border_last

def num_format(sheet, start_row, end_row, start_col, end_col):
    Dollar_keys = ['amount', 'revenue', 'spread', 'amt', 'spl', 'avg loss', 'loss per load', 'losses', 'cost', '$']
    percent_keys = ['percent', 'pct', 'percentage', '%']
    date_keys = ['date']

    dollar_f = '$#,##0.00'
    percent_f = '##0.00%'
    num_f = '#,##0'
    date = 'mm-dd-yy'

    for row in range(start_row + 1, end_row + 1):
        for col in range(start_col, end_col + 1):
            sheet.cell(row=row, column=col).number_format = num_f

    for col in range(start_col, end_col + 1):
        if sheet.cell(row=start_row, column=col).value is not None:

            for dtk in date_keys:
                if dtk in sheet.cell(row=start_row, column=col).value.lower():
                    for row in range(start_row + 1, end_row + 1):
                        sheet.cell(row=row, column=col).number_format = date

            for dk in Dollar_keys:
                if dk in sheet.cell(row=start_row, column=col).value.lower():
                    for row in range(start_row + 1, end_row + 1):
                        sheet.cell(row=row, column=col).number_format = dollar_f

            for pk in percent_keys:
                if pk in sheet.cell(row=start_row, column=col).value.lower():
                    for row in range(start_row + 1, end_row + 1):
                        sheet.cell(row=row, column=col).number_format = percent_f

        else:
            pass


def red_cell(sheet, row, col):
    fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
    font = Font(color = 'D32F2F', bold = True) #, size = 20)

    sheet.cell(row = row, column = col).fill = fill
    sheet.cell(row = row, column = col).font = font

def green_cell(sheet, row, col):
    fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
    font = Font(color='388E3C', bold=True)  # , size = 20)

    sheet.cell(row=row, column=col).fill = fill
    sheet.cell(row=row, column=col).font = font

def total_cell(sheet, row, col):
    fill = PatternFill(start_color='455A64', end_color='455A64', fill_type='solid')
    font = Font(color='FFFFFF', bold=True, italic=True)  # , size = 20)

    sheet.cell(row=row, column=col).fill = fill
    sheet.cell(row=row, column=col).font = font

def conditional_formatting(sheet, n, data, today, start_row, end_row, start_col, end_col):
    for col in range(start_col, end_col + 1):
        if sheet.cell(row=start_row, column=col).value is not None:
            if '%' in sheet.cell(row=start_row, column=col).value.lower() and (
                    'loss' not in sheet.cell(row=start_row, column=col).value.lower()
                    and 'cost' not in sheet.cell(row=start_row, column=col).value.lower()
                    and 'roll' not in sheet.cell(row=start_row, column=col).value.lower()
                    and 'bounce' not in sheet.cell(row=start_row, column=col).value.lower()
            ) and (
                    'deviation' in sheet.cell(row=start_row, column=col).value.lower()
                    or 'yoy change %' in sheet.cell(row=start_row, column=col).value.lower()
            ):
                if len(data) > n:
                    x = n
                else:
                    x = len(data)
                for row in range(start_row + 1, start_row + x + 1):
                    try:
                        if float(sheet.cell(row=row, column=col).value) < 0:
                            red_cell(sheet, row, col)
                        else:
                            green_cell(sheet, row, col)
                    except:
                        pass

            elif '%' in sheet.cell(row=start_row, column=col).value.lower() and (
                    'loss' in sheet.cell(row=start_row, column=col).value.lower()
                    or 'cost' in sheet.cell(row=start_row, column=col).value.lower()
                    or 'bounce' in sheet.cell(row=start_row, column=col).value.lower()
                    or 'roll' in sheet.cell(row=start_row, column=col).value.lower()
            ) and (
                    'deviation' in sheet.cell(row=start_row, column=col).value.lower()
                    or 'yoy change %' in sheet.cell(row=start_row, column=col).value.lower()
            ):
                if len(data) > n:
                    x = n
                else:
                    x = len(data)
                for row in range(start_row + 1, start_row + x + 1):
                    try:
                        if float(sheet.cell(row=row, column=col).value) > 0:
                            red_cell(sheet, row, col)
                        else:
                            green_cell(sheet, row, col)
                    except:
                        pass

        else:
            pass

def totals(sheet, n, data, today, start_row, end_row, start_col, end_col):
    for row in range(start_row+1, end_row + 1):
        if isinstance(sheet.cell(row=row, column=start_col).value, str):
            if sheet.cell(row=row, column=start_col).value.lower() == 'totals':
                for col in range(start_col, end_col + 1):
                    total_cell(sheet, row, col)